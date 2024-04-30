
from config import DB_CONFIG
from openpyxl import load_workbook
import mysql.connector
import os

# Function to fetch data from MySQL
def fetch_data(query):
    conn = mysql.connector.connect(**DB_CONFIG)
    cursor = conn.cursor()
    cursor.execute(query)
    data = cursor.fetchone()
    conn.close()
    return data





# Function to fill Risk_Assessment.xlsx
def fill_risk_assessment(company_info):
    # Fetch data from MySQL tables
    risk_assessment = fetch_data("SELECT Q1_note, Q2, Q3, Q3_note, Q4, Q4_note, Q5_note, Q6, Q6_note, Q7, Q7_note, Q8, Q8_note, total_risk FROM risk_assessment")
    trade_compliance = fetch_data("SELECT Q9, Q10, Q11, Q12, Q13, Q14, total_trade_compliance FROM trade_compliance")

    # Load Risk_Assessment.xlsx template
    wb = load_workbook('C:/xampp/htdocs/sales-process/python/templates/Risk_Assessment.xlsx')
    ws = wb.active

    # Fill cells with data
    ws['D5'] = company_info[1]  # company_name
    ws['D6'] = f"{company_info[2]}, {company_info[3]}, {company_info[4]}, {company_info[5]}"  # company_address_street, city, state, zipcode
    ws['D10'] = risk_assessment[0]  # Q1_note
    ws['C11'] = risk_assessment[1]  # Q2
    ws['C12'] = risk_assessment[2]  # Q3
    ws['D12'] = risk_assessment[3]  # Q3_note
    ws['C13'] = risk_assessment[4]  # Q4
    ws['D13'] = risk_assessment[5]  # Q4_note
    ws['D14'] = risk_assessment[6]  # Q5_note
    ws['C15'] = risk_assessment[7]  # Q6
    ws['D15'] = risk_assessment[8]  # Q6_note
    ws['C16'] = risk_assessment[9]  # Q7
    ws['D16'] = risk_assessment[10]  # Q7_note
    ws['C18'] = risk_assessment[11]  # Q8
    ws['D18'] = risk_assessment[12]  # Q8_note
    ws['C20'] = risk_assessment[13]  # total_risk
    ws['C28'] = trade_compliance[0]  # Q9
    ws['C29'] = trade_compliance[1]  # Q10
    ws['C30'] = trade_compliance[2]  # Q11
    ws['C31'] = trade_compliance[3]  # Q12
    ws['C32'] = trade_compliance[4]  # Q13
    ws['C34'] = trade_compliance[5]  # Q14
    ws['C36'] = trade_compliance[6]  # total_trade_compliance

    return wb

# Function to fill Customer_Code.xlsx
def fill_customer_code(company_info):
    # Fetch data from MySQL tables
    customer_code = fetch_data("SELECT submitted_by, submission_date, sales_tax, payment_term, delivery_term, Q1_code, Q2_code FROM customer_code")

    # Load Customer_Code.xlsx template
    wb = load_workbook('C:/xampp/htdocs/sales-process/python/templates/Customer_Code.xlsx')
    ws = wb.active

    # Fill cells with data
    ws['E5'] = customer_code[0]  # submitted_by
    ws['M5'] = customer_code[1]  # submission_date
    ws['E6'] = customer_code[0]  # submitted_by
    ws['A22'] = company_info[1]  # company_name
    ws['A23'] = company_info[2]  # company_address_street
    ws['A24'] = f"{company_info[3]}, {company_info[4]} "  # company_address_city, state
    ws['A25'] = company_info[5] #zipcode

    # Determine if shipping and billing addresses are the same
    if (company_info[2], company_info[3], company_info[4], company_info[5]) != (company_info[10], company_info[11], company_info[12], company_info[13]):
        ws['I22'] = company_info[1]  # company_name
        ws['I23'] = company_info[10]  # shipping_address_street
        ws['I24'] = f"{company_info[11]} {company_info[12]} " # shipping_address_city, state,
        ws['I25'] =  company_info[13] # zipcode
    else:
        ws['I22'] = "THE SAME"

    # tax, payment term, delivery.
    ws['E28'] = customer_code[2]  # sales_tax
    ws['M27'] = customer_code[3]  # payment_term
    ws['M28'] = customer_code[4]  # delivery_term
    ws['O36'] = customer_code[5]  # Q1_code
    ws['O40'] = customer_code[6]  # Q2_code

    return wb

# Main function
def main():
    # Fetch company info from MySQL
    company_info = fetch_data("SELECT * FROM company_info")

    # Define the folder path where the Excel files will be saved
    folder_path = fr'C:\Users\Rettis\OneDrive - 株式会社レゾナック\Customer Approval Repository\{company_info[1]}'

    # Check if the folder already exists
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)

    # Fill Risk_Assessment.xlsx and save
    risk_assessment_wb = fill_risk_assessment(company_info)
    risk_assessment_file_path = os.path.join(folder_path, f'{company_info[1]}_Risk_Assessment.xlsx')
    risk_assessment_wb.save(risk_assessment_file_path)

    # Fill Customer_Code.xlsx and save
    customer_code_wb = fill_customer_code(company_info)
    customer_code_file_path = os.path.join(folder_path, f'{company_info[1]}_Customer_Code.xlsx')
    customer_code_wb.save(customer_code_file_path)

if __name__ == "__main__":
    main()
